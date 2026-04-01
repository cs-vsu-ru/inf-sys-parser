from copy import copy
from typing import TypeAlias

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.employees.models import Employee
from app.lessons.models import Lesson

TargetFilename: TypeAlias = str

_COLS_PER_TEACHER = 2
_FIXED_LEFT_COLS = 2

_HEADER_FONT = Font(bold=True, size=12)
_CELL_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

_COLUMN_WIDTH = 17
_HEADER_ROW_HEIGHT_MIN_PT = 18
_ROW_PT_PER_TEXT_LINE = 16.0

_NO_FILL = PatternFill(fill_type=None)


def _pair_width_chars() -> int:
    """Эквивалентная ширина текста для пары объединённых столбцов (для переноса)."""
    return max(8, int(_COLUMN_WIDTH * _COLS_PER_TEACHER * 0.85))


def _estimate_row_height_pt(text: str | None, width_chars: int) -> float:
    """Грубая оценка высоты строки в пунктах под перенос по ширине width_chars."""
    if not text:
        return _ROW_PT_PER_TEXT_LINE
    s = str(text).replace('\r', '')
    lines = 0
    cur = 0
    wc = max(4, width_chars)
    for ch in s:
        if ch == '\n':
            lines += 1
            cur = 0
        else:
            cur += 1
            if cur > wc:
                lines += 1
                cur = 1
    lines += 1
    h = 18.0 * lines + 10.0
    return min(409.0, max(_ROW_PT_PER_TEXT_LINE, h))


def _last_data_column(num_employees: int) -> int:
    """Последняя колонка с данными преподавателей (пара столбцов на человека)."""
    return _FIXED_LEFT_COLS + num_employees * _COLS_PER_TEACHER


def _try_merge(ws, min_r: int, min_c: int, max_r: int, max_c: int) -> None:
    """Merge range; ignore if the sheet already has an equivalent merge (template)."""
    try:
        ws.merge_cells(
            start_row=min_r,
            start_column=min_c,
            end_row=max_r,
            end_column=max_c,
        )
    except ValueError:
        pass


def _merge_header_fio_pair(ws, left: int, right: int) -> None:
    """Склеить две ячейки строки 1 под ФИО; убрать конфликтующий merge из шаблона."""
    ws.cell(row=1, column=right).value = None
    coord = f'{get_column_letter(left)}1:{get_column_letter(right)}1'
    try:
        ws.unmerge_cells(coord)
    except ValueError:
        pass
    _try_merge(ws, 1, left, 1, right)


class XlsxEmployeesScheduleBuilder:
    def __init__(self):
        self.template_filename = 'employees_schedule_template.xlsx'
        self.target_filename = 'employees_schedule.xlsx'
        self.lesson_manager = Lesson.objects
        self.employees_manager = Employee.objects

    def build(self) -> TargetFilename:
        employees = list(self.employees_manager.order_by('name'))
        lessons = self.lesson_manager.index_lessons(self.lesson_manager.all())
        wb = load_workbook(self.template_filename)
        ws = wb.active
        template_right_col = ws.max_column
        max_slots = max(
            0, (template_right_col - _FIXED_LEFT_COLS) // _COLS_PER_TEACHER
        )
        extra = len(employees) - max_slots
        if extra > 0:
            ws.insert_cols(template_right_col + 1, extra * _COLS_PER_TEACHER)

        for index, employee in enumerate(employees):
            left = self._teacher_left_col(index, max_slots, template_right_col)
            right = left + 1
            _merge_header_fio_pair(ws, left, right)
            hdr = ws.cell(row=1, column=left)
            hdr.value = employee.name

            for weekday in range(6):
                for number in range(8):
                    numerator_row = self._get_numerator_row(weekday, number)
                    denominator_row = numerator_row + 1
                    numerator_lesson = lessons[(employee, weekday, number, False)]
                    denominator_lesson = lessons[(employee, weekday, number, True)]
                    if self._check_merge(numerator_lesson, denominator_lesson):
                        _try_merge(ws, numerator_row, left, denominator_row, right)
                        cl = ws.cell(row=numerator_row, column=left)
                        cl.value = self._get_cell(numerator_lesson)
                        cl.alignment = _CELL_ALIGNMENT
                    else:
                        _try_merge(ws, numerator_row, left, numerator_row, right)
                        cn = ws.cell(row=numerator_row, column=left)
                        cn.value = self._get_cell(numerator_lesson)
                        cn.alignment = _CELL_ALIGNMENT
                        _try_merge(ws, denominator_row, left, denominator_row, right)
                        cd = ws.cell(row=denominator_row, column=left)
                        cd.value = self._get_cell(denominator_lesson)
                        cd.alignment = _CELL_ALIGNMENT

        max_r = self._get_numerator_row(5, 7) + 1
        last_col = _last_data_column(len(employees))

        self._trim_trailing_columns(ws, last_col)

        self._apply_center_alignment_all_schedule_cells(
            ws, employees, max_slots, template_right_col, max_r, last_col
        )

        if extra > 0:
            anchor_left = template_right_col - 1
            anchor_right = template_right_col
            for pair_i in range(extra):
                new_left = template_right_col + 1 + pair_i * _COLS_PER_TEACHER
                new_right = new_left + 1
                for row in range(1, max_r + 1):
                    src_l = ws.cell(row=row, column=anchor_left)
                    src_r = ws.cell(row=row, column=anchor_right)
                    dst_l = ws.cell(row=row, column=new_left)
                    dst_r = ws.cell(row=row, column=new_right)
                    dst_l.border = copy(src_l.border)
                    dst_r.border = copy(src_r.border)

        self._sync_borders_and_clear_fills(
            ws, max_r, last_col, employees, max_slots, template_right_col
        )
        self._apply_column_widths(ws, last_col)
        self._apply_dynamic_row_heights(ws, employees, lessons, max_slots, template_right_col, max_r)

        wb.save(self.target_filename)
        return self.target_filename

    def _trim_trailing_columns(self, ws, last_data_col: int) -> None:
        """Убрать хвостовые столбцы шаблона (пустой AA и т.п.), чтобы не тянуть лишнюю сетку."""
        excess = ws.max_column - last_data_col
        if excess > 0:
            ws.delete_cols(last_data_col + 1, excess)

    def _apply_center_alignment_all_schedule_cells(
            self,
            ws,
            employees: list,
            max_slots: int,
            template_right_col: int,
            max_r: int,
            last_col: int,
    ) -> None:
        """Центр по горизонтали и вертикали; только реальные колонки данных (без расширения max_column)."""
        for row in range(1, max_r + 1):
            for col in range(1, last_col + 1):
                ws.cell(row=row, column=col).alignment = _CELL_ALIGNMENT
        for index in range(len(employees)):
            left = self._teacher_left_col(index, max_slots, template_right_col)
            ws.cell(row=1, column=left).font = _HEADER_FONT

    def _sync_borders_and_clear_fills(
            self,
            ws,
            max_r: int,
            last_col: int,
            employees: list,
            max_slots: int,
            template_right_col: int,
    ) -> None:
        """Сброс заливки (артефакты «половинок»); в паре столбцов преподавателя — одинаковая рамка."""
        for row in range(1, max_r + 1):
            for col in range(1, last_col + 1):
                ws.cell(row=row, column=col).fill = _NO_FILL
        for index in range(len(employees)):
            left = self._teacher_left_col(index, max_slots, template_right_col)
            right = left + 1
            for row in range(1, max_r + 1):
                cl = ws.cell(row=row, column=left)
                cr = ws.cell(row=row, column=right)
                cr.border = copy(cl.border)

    def _apply_column_widths(self, ws, last_col: int) -> None:
        for col in range(1, last_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = _COLUMN_WIDTH

    def _apply_dynamic_row_heights(
            self,
            ws,
            employees: list,
            lessons,
            max_slots: int,
            template_right_col: int,
            max_r: int,
    ) -> None:
        wc = _pair_width_chars()
        h1 = float(_HEADER_ROW_HEIGHT_MIN_PT)
        for emp in employees:
            h1 = max(h1, _estimate_row_height_pt(emp.name, wc))
        ws.row_dimensions[1].height = h1

        row_need: dict[int, float] = {r: 0.0 for r in range(2, max_r + 1)}

        for employee in employees:
            for weekday in range(6):
                for number in range(8):
                    num_r = self._get_numerator_row(weekday, number)
                    den_r = num_r + 1
                    nl = lessons[(employee, weekday, number, False)]
                    dl = lessons[(employee, weekday, number, True)]
                    t_num = self._get_cell(nl)
                    t_den = self._get_cell(dl)
                    if self._check_merge(nl, dl):
                        block = _estimate_row_height_pt(t_num, wc)
                        half = max(_ROW_PT_PER_TEXT_LINE, block / 2.0)
                        row_need[num_r] = max(row_need[num_r], half)
                        row_need[den_r] = max(row_need[den_r], half)
                    else:
                        row_need[num_r] = max(
                            row_need[num_r], _estimate_row_height_pt(t_num, wc)
                        )
                        row_need[den_r] = max(
                            row_need[den_r], _estimate_row_height_pt(t_den, wc)
                        )

        for col in range(1, _FIXED_LEFT_COLS + 1):
            for row in range(2, max_r + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    w_single = max(4, int(_COLUMN_WIDTH * 0.92))
                    row_need[row] = max(
                        row_need[row],
                        _estimate_row_height_pt(str(cell.value), w_single),
                    )

        for r, h in row_need.items():
            ws.row_dimensions[r].height = min(409.0, max(_ROW_PT_PER_TEXT_LINE, h))

    @staticmethod
    def _teacher_left_col(index: int, max_slots: int, template_right_col: int) -> int:
        """1-based Excel column index of the left cell of a teacher pair."""
        if index < max_slots:
            return _FIXED_LEFT_COLS + 1 + index * _COLS_PER_TEACHER
        return (
                template_right_col
                + 1
                + (index - max_slots) * _COLS_PER_TEACHER
        )

    def _get_numerator_row(self, weekday: int, number: int) -> int:
        return weekday * 16 + number * 2 + 2

    def _get_cell(self, lesson: Lesson) -> str:
        return f"{lesson.name} {lesson.groups} {lesson.placement}"

    def _check_merge(
            self, numerator_lesson: Lesson, denominator_lesson: Lesson
    ) -> bool:
        return (
                numerator_lesson.name == denominator_lesson.name
                and numerator_lesson.groups == denominator_lesson.groups
                and numerator_lesson.placement == denominator_lesson.placement
        )
