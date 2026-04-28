import re
from dataclasses import dataclass, field
from typing import Any

from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from openpyxl import load_workbook

from app.students.models.student import Student, User

COLUMN_ALIASES = {
    'имя': 'first_name',
    'фамилия': 'last_name',
    'логин': 'login',
    'адресэлектроннойпочты': 'email',
    'email': 'email',
    'почта': 'email',
    'группы': 'group',
    'группа': 'group',
}


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    errors: list[dict[str, Any]] = field(default_factory=list)


class StudentsXlsxImporter:
    required_fields = {'first_name', 'last_name', 'login', 'email', 'group'}

    def import_file(self, file) -> ImportResult:
        result = ImportResult()

        workbook = load_workbook(file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)

        try:
            headers_row = next(rows)
        except StopIteration:
            result.errors.append({
                'row': 1,
                'field': 'file',
                'message': 'Файл пустой',
            })
            return result

        header_map = self._build_header_map(headers_row)

        missing = self.required_fields - set(header_map.values())
        if missing:
            result.errors.append({
                'row': 1,
                'field': 'headers',
                'message': 'Отсутствуют обязательные колонки: ' + ', '.join(sorted(missing)),
            })
            return result

        for row_number, row in enumerate(rows, start=2):
            raw_data = self._row_to_data(header_map, row)

            if self._is_empty_row(raw_data):
                continue

            try:
                data = self._normalize_row(raw_data)

                user, user_created = User.objects.update_or_create(
                    login=data['login'],
                    defaults={
                        'first_name': data['first_name'],
                        'last_name': data['last_name'],
                        'email': data['email'],
                        'role': 'USER',
                    },
                )

                student, student_created = Student.objects.update_or_create(
                    user=user,
                    defaults={
                        'group': data['group'],
                    },
                )

                if user_created or student_created:
                    result.created += 1
                else:
                    result.updated += 1

            except Exception as exc:
                result.errors.append({
                    'row': row_number,
                    'field': 'row',
                    'message': str(exc),
                })

        return result

    def _build_header_map(self, headers_row) -> dict[int, str]:
        result = {}

        for index, value in enumerate(headers_row):
            normalized = self._normalize_header(value)
            field = COLUMN_ALIASES.get(normalized)

            if field:
                result[index] = field

        return result

    def _normalize_header(self, value: Any) -> str:
        value = '' if value is None else str(value)
        value = re.sub(r'<[^>]*>', '', value)  # убирает <br/>, <br>, любые html-теги
        value = value.lower()
        value = re.sub(r'[^a-zа-яё0-9]+', '', value)
        return value

    def _row_to_data(self, header_map: dict[int, str], row) -> dict[str, Any]:
        data = {}

        for index, field in header_map.items():
            data[field] = row[index] if index < len(row) else None

        return data

    def _is_empty_row(self, row_data: dict[str, Any]) -> bool:
        return all(
            value is None or str(value).strip() == ''
            for value in row_data.values()
        )

    def _normalize_row(self, row_data: dict[str, Any]) -> dict[str, str]:
        data = {}

        labels = {
            'first_name': 'Имя',
            'last_name': 'Фамилия',
            'login': 'Логин',
            'email': 'Адрес электронной почты',
            'group': 'Группы',
        }

        for field, label in labels.items():
            value = row_data.get(field)

            if value is None or str(value).strip() == '':
                raise ValueError(f'Поле "{label}" обязательно')

            data[field] = str(value).strip()

        data['login'] = data['login'].lower()
        data['email'] = data['email'].lower()

        try:
            validate_email(data['email'])
        except ValidationError as exc:
            raise ValueError('Некорректный Email') from exc

        return data
